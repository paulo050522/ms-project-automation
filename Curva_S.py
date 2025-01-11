import win32com.client
import pandas as pd
from dateutil.parser import parse
from datetime import timedelta
from openpyxl import load_workbook
from openpyxl.styles import Alignment, NamedStyle


def process_project_with_iterations_and_formatting(project_file, output_excel):
    try:
        # Abrir o MS Project
        app = win32com.client.Dispatch("MSProject.Application")
        app.Visible = True

        # Abrir o arquivo do projeto
        app.FileOpen(project_file)
        project = app.ActiveProject

        # Obter a "linha zero" (Project Summary Task)
        summary_task = project.ProjectSummaryTask

        # Verificar se a "linha zero" possui datas de linha de base configuradas
        if not summary_task.BaselineStart or not summary_task.BaselineFinish:
            raise Exception(
                "A linha zero não possui datas de linha de base configuradas.")

        # Obter a data de início e término da linha de base
        baseline_start_date = parse(
            str(summary_task.BaselineStart)).replace(tzinfo=None)
        baseline_finish_date = parse(
            str(summary_task.BaselineFinish)).replace(tzinfo=None)

        # Inicializar lista de resultados
        results = []

        # Iterar a partir da data de início até a data de término, avançando 30 dias
        current_date = baseline_start_date
        while current_date <= baseline_finish_date:
            # Definir a data atual como StatusDate
            project.StatusDate = current_date.strftime("%d/%m/%Y")
            print(
                f"Definindo StatusDate como: {current_date.strftime('%d/%m/%Y')}")

            # Copiar o valor do campo Texto15
            text15_value = summary_task.Text15 if summary_task.Text15 else "0%"

            # Adicionar os resultados à lista
            results.append({
                "StatusDate": current_date.strftime("%Y-%m-%d"),
                "Text15Value": text15_value
            })

            # Avançar para o próximo mês (aproximadamente 30 dias)
            current_date += timedelta(days=30)

        # Salvar o arquivo do MS Project
        app.FileSave()
        print(f"Arquivo do MS Project salvo: {project_file}")

        # Criar um DataFrame com os resultados
        df = pd.DataFrame(results)

        # Converter os valores da coluna Text15Value para números decimais
        df["Text15Value"] = df["Text15Value"].str.replace(
            "%", "").astype(float) / 100

        # Salvar o DataFrame ajustado no Excel
        df.to_excel(output_excel, index=False)
        print(f"Resultados salvos no Excel: {output_excel}")

        # Ajustar a formatação do Excel e mudar para porcentagem inteira
        adjust_excel_formatting(output_excel)

        # Fechar o arquivo do projeto
        app.FileClose(False)
        app.Quit()

    except Exception as e:
        print(f"Erro: {e}")
        if 'app' in locals():
            app.Quit()


def adjust_excel_formatting(output_excel):
    try:
        # Carregar o arquivo Excel gerado
        wb = load_workbook(output_excel)
        ws = wb.active

        # Definir a largura das colunas
        ws.column_dimensions["A"].width = 13.44  # Largura para a coluna A
        ws.column_dimensions["B"].width = 9.56   # Largura para a coluna B

        # Ajustar o alinhamento horizontal para "center" nas colunas A e B
        for row in ws.iter_rows(min_col=1, max_col=2, min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

        # Alterar o formato da coluna B para porcentagem inteira
        percentage_style = NamedStyle(name="percentage_integer")
        percentage_style.number_format = "0%"
        for cell in ws["B"]:
            if cell.row > 1:  # Ignorar o cabeçalho
                cell.style = percentage_style

        # Salvar as alterações no mesmo arquivo
        wb.save(output_excel)
        print(
            f"Formatação ajustada e coluna B configurada como porcentagem inteira no arquivo Excel: {output_excel}")

    except Exception as e:
        print(f"Erro ao ajustar a formatação: {e}")


# Caminho do arquivo do MS Project
project_file_path = r"C:\Users\Petroeng\Documents\Projetos VS CODE\Projetos_Code\6. Project\TESTE.mpp"

# Caminho do arquivo Excel para salvar os dados
output_excel_path = r"C:\Users\Petroeng\Documents\Projetos VS CODE\Projetos_Code\6. Project\Resultado_Final_Formatado.xlsx"

# Executar a função
process_project_with_iterations_and_formatting(
    project_file_path, output_excel_path)
